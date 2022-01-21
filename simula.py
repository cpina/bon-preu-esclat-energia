#!/usr/bin/env python3

import argparse
import os
from datetime import datetime

import openpyxl

acumulats = {0.079990:0, 0.129990:0,0.229990:0 }

def calcular_fesho_facil(data, hora, activa_entrant):
    return activa_entrant * 0.179990 / 1000


def calcular_preu_dia_i_nit(data, hora, activa_entrant):
    global acumulats

    hora = int(hora)

    vall = 0.079990
    plana = 0.129990
    punta = 0.229990

    preus = {1: vall, 2: vall, 3: vall, 4: vall, 5: vall, 6: vall, 7: vall, 8: vall, 9: plana, 10: plana, 11: punta,
             12: punta, 13: punta, 14: punta, 15: plana, 16: plana, 17: plana, 18: plana, 19: punta, 20: punta, 21: punta, 22: punta,
             23: plana, 24: plana
             }

    dt = datetime.strptime(data, "%d/%m/%Y")
    dia_de_la_setmana = dt.weekday()

    if data == "06/12/2021" or data == "08/12/2021":
        preu = vall
    elif dia_de_la_setmana == 5 or dia_de_la_setmana == 6:
        preu = vall
    else:
        preu = preus[hora]

    acumulats[preu] += activa_entrant

    return preu * activa_entrant / 1000


def calcula(fitxer, calcula_preu_func):
    workbook = openpyxl.load_workbook(fitxer)
    sheet = workbook.get_active_sheet()

    row = 2

    preu_acumulat = 0
    activa_entrant_acumulat = 0

    while True:
        cups = sheet.cell(row=row, column=1).value
        if cups is None:
            break

        data = sheet.cell(row=row, column=2).value
        hora = sheet.cell(row=row, column=3).value
        activa_entrant = sheet.cell(row=row, column=4).value
        preu_acumulat += calcula_preu_func(data, hora, activa_entrant)

        row += 1

    return preu_acumulat


def simula(fitxer):
    global acumulats

    preu_dia_i_nit = calcula(fitxer, calcular_preu_dia_i_nit)
    fesho_facil = calcula(fitxer, calcular_fesho_facil)
    print(os.path.basename(fitxer))
    print(f"Dia i nit   : {preu_dia_i_nit:.2f}€")
    print(f"Fes-ho fàcil: {fesho_facil:.2f}€")

    for preu, energia in acumulats.items():
        print(preu, energia/1000)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()

    parser.add_argument('fitxer', help='Fitxer .xlsx de Bon Preu Esclat en franges horaries')
    args = parser.parse_args()

    simula(args.fitxer)
